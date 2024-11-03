import os
import openpyxl

user_input = input("Matric Number::")
grpA = os.path.join("C:\\Users\\hp\\Desktop\\grpA_culcsc.xlsx")
list_students = []
workBook = openpyxl.load_workbook(grpA)
sheet = workBook[workBook.sheetnames[0]]

length_workBook = sheet.max_row
for each in range(2,length_workBook):
    if sheet.cell(row=each, column=2).value:
        test = {"S/N":sheet.cell(row=each, column=1).value,"name":sheet.cell(row=each, column=2).value.lower(),"matricNumber":sheet.cell(row=each, column=3).value}
        list_students.append(test)
for student in list_students:
    if student["matricNumber"] == user_input:
        print(f"{student["S/N"]} {student["name"].upper()} {student["matricNumber"]}")